"""Migra o XLSM legado para a base privada do OneDrive.

Nenhum dado gerado por este script é enviado ao GitHub.
"""
from __future__ import annotations
import json, hashlib, os, re, shutil, unicodedata, uuid
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(os.environ["CASA_EM_ORDEM_MIGRATION_SOURCE"]) if os.environ.get("CASA_EM_ORDEM_MIGRATION_SOURCE") else next(p for p in (ROOT / "Acompanhamento").glob("*.xlsm") if "backup" not in p.name.lower())
TARGET = Path.home() / "OneDrive" / "CasaEmOrdem-familia.json"
NOW = datetime.now().astimezone().isoformat()

def norm(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9]", " ", "".join(c for c in text if not unicodedata.combining(c)).upper())).strip()

def ident(): return str(uuid.uuid4())
def audit(): return {"id": ident(), "createdAt": NOW, "updatedAt": NOW, "updatedBy": "Ambos", "version": 1}
def iso(value):
    if isinstance(value, (datetime, date)): return value.isoformat()[:10]
    return str(value or "")[:10]

data = json.loads(TARGET.read_text(encoding="utf-8"))
backup = TARGET.with_name(f"CasaEmOrdem-familia-backup-{datetime.now():%Y%m%d-%H%M%S}.json")
shutil.copy2(TARGET, backup)
try:
    wb = load_workbook(SOURCE, read_only=True, data_only=True, keep_vba=True)
except PermissionError:
    # O Excel pode manter o arquivo do OneDrive bloqueado. Uma cópia de leitura
    # preserva o livro aberto e permite a migração sem executá-lo.
    import tempfile
    snapshot = Path(tempfile.gettempdir()) / "CasaEmOrdem-Financas-snapshot.xlsm"
    shutil.copyfile(SOURCE, snapshot)
    wb = load_workbook(snapshot, read_only=True, data_only=True, keep_vba=True)

categories = {norm(c["name"]): c for c in data["categories"]}
def category(name, sub=""):
    key = norm(name) or "OUTROS"
    c = categories.get(key)
    if not c:
        display = str(name or "Outros").strip()
        nature = "income" if key in {"RECEITAS", "EMPRESA"} else "transfer" if "TRANSFER" in key else "expense"
        c = {**audit(), "name": display, "subcategories": [], "nature": nature}
        data["categories"].append(c); categories[key] = c
    if sub and sub not in c["subcategories"]: c["subcategories"].append(sub)
    return c

accounts = {}
def account(origin):
    key = norm(origin)
    if key not in accounts:
        operator = "Mari" if "MARI" in key else "Olcino" if " OL" in f" {key}" else "Ambos"
        institution = next((x for x in ("Inter", "XP", "BTG", "Mercado Pago") if norm(x) in key), "Outro")
        kind = "card" if "CARTAO" in key else "checking"
        obj = {**audit(), "name": str(origin), "institution": institution, "kind": kind, "operator": operator, "active": True}
        data["accounts"].append(obj); accounts[key] = obj
    return accounts[key]

# O arquivo ainda estava vazio; a migração é recriada de forma determinística.
data["accounts"] = []; data["transactions"] = []; data["rules"] = []; data["imports"] = []
existing_accounts = accounts
ws = wb[wb.sheetnames[1]]
headers = [str(v or "").strip() for v in next(ws.iter_rows(values_only=True))]
idx = {norm(v): i for i, v in enumerate(headers)}
def col(row, *names):
    for name in names:
        i = idx.get(norm(name))
        if i is not None and i < len(row): return row[i]
    return None

seen = set(); rule_counts = Counter(); ignored = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    when = col(row, "Data ajustada", "Data")
    description = col(row, "Lancamento")
    value = col(row, "Valor")
    origin = col(row, "Origem")
    if not when or not description or not origin or not isinstance(value, (int, float)): continue
    cat_text = str(col(row, "SUBCATEGORIA - CONTA (Definida)") or col(row, "SUBCATEGORIA - CONTA (Sugerida)") or "")
    movement = norm(col(row, "MOVIMENTAÇÃO"))
    cat_name, _, sub = cat_text.partition("-")
    cat = category(cat_name or "Outros", sub or "Compras diversas")
    acc = account(origin)
    d = iso(when); installment = col(row, "Parcela.1")
    key_raw = f"{d}|{norm(description)}|{float(value):.2f}|{acc['id']}|{installment or 0}"
    key = hashlib.sha256(key_raw.encode()).hexdigest()
    if key in seen: ignored += 1; continue
    seen.add(key)
    transfer = "TRANSFERENCIAENTRECONTAS" in norm(cat_text) or "TRANSFER" in movement or "PAGAMENTO ON LINE" in norm(description) or "FATURA CARTAO" in norm(description)
    movement_kind = "transfer" if transfer else "reserve" if "RESERVA" in movement else "expense_income"
    source_kind = "card" if "CARTAO" in norm(origin) else "statement"
    is_income = cat["nature"] == "income" or "RECEITA" in norm(sub)
    amount = -abs(float(value)) if is_income else abs(float(value))
    operator = acc["operator"]
    scope = "Transferência interna" if transfer else "Familiar" if operator == "Ambos" else f"Pessoal — {operator}"
    t = {**audit(), "date": d, "competence": d[:7], "description": str(description), "normalized": norm(description), "amount": amount, "accountId": acc["id"], "operator": operator, "scope": scope, "categoryId": cat["id"], "subcategory": sub or "Compras diversas", "classification": "confirmed" if cat_text else "pending", "dedupeKey": key, "transfer": transfer, "movement": movement_kind, "sourceKind": source_kind}
    if isinstance(installment, (int, float)) and installment > 0: t["installment"] = int(installment)
    data["transactions"].append(t)
    if cat_text: rule_counts[(t["normalized"], acc["id"], operator, cat["id"], t["subcategory"])] += 1

for (pattern, account_id, operator, category_id, subcategory), hits in rule_counts.items():
    data["rules"].append({**audit(), "pattern": pattern, "match": "exact", "categoryId": category_id, "subcategory": subcategory, "accountId": account_id, "operator": operator, "priority": 100, "active": True, "hits": hits})

# Orçamento vigente e central de pagamentos a partir do planejamento mensal.
data["budgets"] = [b for b in data["budgets"] if b.get("member")]
data["obligations"] = []
monthly = wb[wb.sheetnames[6]]
for row in monthly.iter_rows(min_row=17, values_only=True):
    cat_name, name, value = row[1], row[2], row[3]
    if not cat_name or not name or not isinstance(value, (int, float)) or not value: continue
    is_income = value > 0 and "RECEITA" in norm(name)
    cat = category("Receitas" if is_income else cat_name, str(name))
    if not is_income: data["budgets"].append({**audit(), "month": "2026-07", "startMonth": "2026-07", "categoryId": cat["id"], "amount": abs(float(value)), "reason": f"Migrado de {name}"})
    control = str(row[8] or "") if len(row) > 8 else ""
    due_day = int(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else 1
    if not is_income and "PAGAR" in norm(control):
        kind = "Recorrência no cartão" if "CARTAO" in norm(control) else "Manual"
        data["obligations"].append({**audit(), "name": str(name), "kind": kind, "planned": abs(float(value)), "dueDate": f"2026-07-{min(due_day,28):02d}", "recurrence": "monthly", "tolerance": max(1, abs(float(value))*.05), "status": "A pagar"})

# Reservas de despesas e objetivos de longo prazo.
data["goals"] = []
for sheet_index, goal_kind in ((7, "Reserva de despesa"), (8, "Objetivo")):
    sheet = wb[wb.sheetnames[sheet_index]]
    for row in sheet.iter_rows(min_row=17, values_only=True):
        cat_name, name, monthly_value = row[1], row[2], row[3]
        if not name or not isinstance(monthly_value, (int, float)): continue
        target = row[6] if sheet_index == 7 else row[4]
        current = row[7] if sheet_index == 7 else row[5]
        deadline = row[9] if sheet_index == 7 else row[7]
        if not isinstance(target, (int, float)) or target <= 0: continue
        goal = {**audit(), "name": str(name), "target": abs(float(target)), "deadline": iso(deadline) or "2030-12-31", "priority": len(data["goals"])+1, "minimum": abs(float(monthly_value)), "emergency": "EMERGENCIA" in norm(name), "active": True, "movements": []}
        if isinstance(current, (int, float)) and current: goal["movements"].append({"id": ident(), "date": "2026-07-01", "kind": "ajuste", "amount": float(current), "reason": "Saldo migrado da planilha"})
        data["goals"].append(goal)

data["imports"] = [{**audit(), "filename": SOURCE.name, "hash": hashlib.sha256(SOURCE.read_bytes()).hexdigest(), "institution": "Histórico consolidado", "count": len(data["transactions"]), "duplicates": ignored}]
data["lastSavedAt"] = NOW
TARGET.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({"backup": str(backup), "transactions": len(data["transactions"]), "rules": len(data["rules"]), "categories": len(data["categories"]), "budgets": len(data["budgets"]), "obligations": len(data["obligations"]), "goals": len(data["goals"]), "duplicates_ignored": ignored}, ensure_ascii=False, indent=2))
